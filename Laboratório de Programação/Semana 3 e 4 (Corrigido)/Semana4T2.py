import openpyxl
import json
import random
import sys
from typing import Dict, List, Tuple


# --- Constantes de configuração ---
FICHEIRO_XLSX: str = "tabela_sinoptica.xlsx"
FICHEIRO_NOMES: str = "nomes.txt"
FICHEIRO_APELIDOS: str = "apelidos.txt"
FICHEIRO_SAIDA_JSON: str = "receitas_medicas.json"

NUM_UTENTES: int = 10
MIN_MEDICAMENTOS_RECEITA: int = 2
MAX_MEDICAMENTOS_RECEITA: int = 6
LIMIAR_INTERACAO: int = 1

LEGENDA_INTERACOES: Dict[int, str] = {
    0: "Sem significado clínico",
    1: "Potencialmente grave",
    2: "Potenciador do efeito terapêutico/tóxico (coluna horizontal)",
    3: "Potenciador do efeito terapêutico/tóxico (coluna vertical)",
    4: "Diminuidor do efeito terapêutico/tóxico (coluna horizontal)",
    5: "Diminuidor do efeito terapêutico/tóxico (coluna vertical)",
}


# --- Tipos auxiliares ---
TabelaSinoptica = Dict[str, Dict[str, int]]
Utente = Dict[str, str]
Receita = Dict[str, List[str]]

def normalizar_nome(nome: str) -> str:
    """
    Normalização simplificada: apenas minúsculas e sem espaços extra.

    :param nome: Nome a normalizar.
    :return: Nome normalizado.
    """
    if not nome: 
        return ""
    return str(nome).strip().lower()


# --- Funções de leitura ---
def ler_lista_ficheiro(caminho: str) -> List[str]:
    """
    Lê uma lista de um ficheiro de texto, ignorando linhas vazias.

    :param caminho: Caminho para o ficheiro.
    :return: Lista de strings lidas do ficheiro.
    """
    try:
        with open(caminho, "r", encoding="utf-8") as f:
            return [linha.strip() for linha in f if linha.strip()]
    except FileNotFoundError:
        print(f"Erro: Ficheiro '{caminho}' não encontrado.")
        sys.exit(1)


def ler_tabela_excel(caminho: str) -> Tuple[List[str], TabelaSinoptica]:
    """
    Lê a tabela do Excel usando openpyxl.

    :param caminho: Caminho para o ficheiro Excel.
    :return: Tupla contendo a lista de nomes e a tabela sinóptica.
    """
    try:
        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = wb.active

        nomes_originais = []
        for col in range(2, ws.max_column + 1):
            valor = ws.cell(row=1, column=col).value
            if valor:
                nomes_originais.append(str(valor))

        tabela = {}
        for row in range(2, ws.max_row + 1):
            nome_linha = ws.cell(row=row, column=1).value
            if not nome_linha: continue
            
            chave_linha = normalizar_nome(nome_linha)
            tabela[chave_linha] = {}

            for col_idx, nome_col in enumerate(nomes_originais, start=2):
                chave_coluna = normalizar_nome(nome_col)
                valor_celula = ws.cell(row=row, column=col_idx).value
                
                try:
                    tabela[chave_linha][chave_coluna] = int(valor_celula) if valor_celula is not None else 0
                except (ValueError, TypeError):
                    tabela[chave_linha][chave_coluna] = 0
        return nomes_originais, tabela
    except Exception as e:
        print(f"Erro ao processar Excel: {e}")
        sys.exit(1)


# --- Lógica de Geração e Cálculo ---
def calcular_balanco(receita: List[str], tabela: TabelaSinoptica) -> Dict:
    """
    Calcula interações usando loops aninhados em vez de combinations.

    :param receita: Lista de medicamentos da receita.
    :param tabela: Tabela sinóptica de interações.
    :return: Dicionário com o balanço de interações.
    """
    interacoes = []
    contagem = {v: 0 for v in range(6)}
    
    # Substituição de itertools.combinations(receita, 2)[cite: 10]
    n = len(receita)
    for i in range(n):
        for j in range(i + 1, n):
            med_a = receita[i]
            med_b = receita[j]
            
            v = tabela.get(normalizar_nome(med_a), {}).get(normalizar_nome(med_b), 0)
            contagem[v] += 1
            interacoes.append({
                "medicamento_a": med_a,
                "medicamento_b": med_b,
                "valor": v,
                "descricao": LEGENDA_INTERACOES.get(v, "Desconhecido")
            })

    tem_interacoes = any(p["valor"] >= LIMIAR_INTERACAO for p in interacoes)
    return {
        "pares_analisados": len(interacoes),
        "interacoes": sorted(interacoes, key=lambda x: x["valor"], reverse=True),
        "contagem_por_tipo": {str(k): v for k, v in contagem.items()},
        "tem_interacoes": tem_interacoes,
        "decisao": "Com interações" if tem_interacoes else "Sem interações"
    }


def main() -> None:
    """
    Função principal que orquestra a leitura dos dados, geração das receitas e cálculo do balanço de interações.
    
    - Lê a tabela sinóptica do ficheiro Excel.
    - Lê os nomes e apelidos dos utentes.
    - Gera receitas aleatórias para um número definido de utentes.
    - Calcula o balanço de interações para cada receita.
    """
    medicamentos, tabela = ler_tabela_excel(FICHEIRO_XLSX)
    nomes = ler_lista_ficheiro(FICHEIRO_NOMES)
    apelidos = ler_lista_ficheiro(FICHEIRO_APELIDOS)

    receitas = []
    for i in range(NUM_UTENTES):
        utente_nome = f"{random.choice(nomes)} {random.choice(apelidos)}"
        num_meds = random.randint(MIN_MEDICAMENTOS_RECEITA, min(MAX_MEDICAMENTOS_RECEITA, len(medicamentos)))
        lista_meds = random.sample(medicamentos, num_meds)
        
        balanco = calcular_balanco(lista_meds, tabela)
        
        receitas.append({
            "utente_id": f"U-{i+1:05d}",
            "utente_nome": utente_nome,
            "medicamentos_prescritos": lista_meds,
            "num_medicamentos": len(lista_meds),
            "balanco_interacoes": balanco
        })

    with open(FICHEIRO_SAIDA_JSON, "w", encoding="utf-8") as f:
        json.dump({"total_utentes": len(receitas), "receitas": receitas}, f, ensure_ascii=False, indent=4)
    
    print(f"[OK] Geradas {len(receitas)} receitas em '{FICHEIRO_SAIDA_JSON}'.")


if __name__ == "__main__":
    main()