import openpyxl
import random
import sys
from typing import List


# --- Constantes ---
FICHEIRO_MEDICAMENTOS: str = "medicamentos.txt"
FICHEIRO_SAIDA_XLSX: str = "tabela_sinoptica.xlsx"
VALOR_MIN: int = 0
VALOR_MAX: int = 5

def ler_medicamentos(caminho_ficheiro: str) -> List[str]:
    """
    Lê a lista de medicamentos ignorando linhas vazias.

    :param caminho_ficheiro: Caminho para o ficheiro de medicamentos.
    :return: Lista de medicamentos.
    """
    try:
        with open(caminho_ficheiro, "r", encoding="utf-8") as ficheiro:
            return [linha.strip() for linha in ficheiro if linha.strip()]
    except FileNotFoundError:
        print(f"Erro: O ficheiro '{caminho_ficheiro}' não foi encontrado.")
        sys.exit(1)


def gerar_tabela_sinoptica(medicamentos: List[str]) -> List[List[int]]:
    """
    Gera a matriz quadrada com 0 na diagonal.

    :param medicamentos: Lista de medicamentos.
    :return: Matriz de interações.
    """
    n = len(medicamentos)
    return [[0 if i == j else random.randint(VALOR_MIN, VALOR_MAX) 
             for j in range(n)] for i in range(n)]


def guardar_excel(medicamentos: List[str], tabela: List[List[int]], caminho_saida: str) -> None:
    """
    Guarda a tabela num ficheiro Excel nativo (.xlsx).

    :param medicamentos: Lista de medicamentos.
    :param tabela: Matriz de interações.
    :param caminho_saida: Caminho para o ficheiro de saída.
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Interações"

        # 1. Escrever cabeçalho (Linha 1, começando na Coluna B)
        for col_idx, nome in enumerate(medicamentos, start=2):
            ws.cell(row=1, column=col_idx, value=nome)

        # 2. Escrever linhas (Nome na Coluna A + Valores)
        for row_idx, nome_med in enumerate(medicamentos, start=2):
            ws.cell(row=row_idx, column=1, value=nome_med)  # Coluna A
            for col_idx, valor in enumerate(tabela[row_idx-2], start=2):
                ws.cell(row=row_idx, column=col_idx, value=valor)

        wb.save(caminho_saida)
        print(f"[OK] Ficheiro Excel '{caminho_saida}' gerado com sucesso.")
    except Exception as e:
        print(f"Erro ao guardar Excel: {e}")
        sys.exit(1)


def main():
    """
    Função principal que orquestra a leitura dos medicamentos, geração da tabela sinóptica e guarda do ficheiro Excel.

    - Lê os medicamentos do ficheiro de texto.
    - Gera a tabela sinóptica com interações aleatórias.
    - Guarda a tabela num ficheiro Excel nativo (.xlsx).
    - Inclui tratamento de erros para ficheiros não encontrados e problemas na escrita do Excel.
    - Mantém a estrutura modular e clara para fácil manutenção e extensão futura.
    """
    medicamentos = ler_medicamentos(FICHEIRO_MEDICAMENTOS)
    if not medicamentos: return
    
    tabela = gerar_tabela_sinoptica(medicamentos)
    guardar_excel(medicamentos, tabela, FICHEIRO_SAIDA_XLSX)


if __name__ == "__main__":
    main()